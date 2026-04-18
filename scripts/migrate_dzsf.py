"""One-shot importer: DZSF raw XLSX → incidents/*.yaml (with needs_review=True)."""
from __future__ import annotations

import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

COUNTRY_MAP = {
    "china": "CN", "deutschland": "DE", "germany": "DE", "russland": "RU",
    "russia": "RU", "ukraine": "UA", "usa": "US", "vereinigte staaten": "US",
    "südkorea": "KR", "nordkorea": "KP", "uk": "GB", "vereinigtes königreich": "GB",
    "großbritannien": "GB", "italien": "IT", "frankreich": "FR", "spanien": "ES",
    "österreich": "AT", "schweiz": "CH", "niederlande": "NL", "belgien": "BE",
    "polen": "PL", "tschechien": "CZ", "dänemark": "DK", "schweden": "SE",
    "norwegen": "NO", "finnland": "FI", "iran": "IR", "israel": "IL",
    "kanada": "CA", "australien": "AU", "indien": "IN", "japan": "JP",
}

VECTOR_MAP = {
    "exploit": "exploit",
    "exploit/backdoor": "exploit",
    "schwachstelle": "exploit",
    "phishing": "phishing",
    "social engineering": "phishing",
    "supply chain": "supply_chain",
    "supply-chain": "supply_chain",
    "insider": "insider",
    "physisch": "physical",
    "credential": "credential_theft",
    "unklar": "unknown",
    "": "unknown",
}

SYSTEM_COLUMNS = [
    ("betr. System unklar", "unclear_system"),
    ("interne Komm.", "internal_communication"),
    ("Komm. mit Fahrzeugführer", "train_driver_communication"),
    ("Bezahlsystem", "payment_system"),
    ("IT in Bahnhöfen", "station_it"),
    ("Internes Zugtracking", "internal_train_tracking"),
    ("Passagier-infos", "passenger_information"),
    ("Webseite", "website"),
    ("Daten", "data"),
    ("Bahnbetrieb", "rail_operations"),
]


def _slugify(raw: str) -> str:
    s = raw.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "unknown"


def _parse_date(raw: str) -> tuple[date, str]:
    """Return (date, precision) from strings like '25.12.2014' or '01.07.2015*'."""
    precision = "day"
    s = raw.strip()
    if s.endswith("*"):
        precision = "month"
        s = s[:-1]
    d = datetime.strptime(s, "%d.%m.%Y").date()
    if precision == "month":
        d = date(d.year, d.month, 1)
    return d, precision


def _countries(raw: str) -> list[str]:
    if not raw:
        return []
    out: list[str] = []
    for part in re.split(r"[,/;]", raw):
        key = part.strip().lower()
        if not key:
            continue
        code = COUNTRY_MAP.get(key)
        if code and code not in out:
            out.append(code)
    return out


def _bool_de(raw: Any) -> bool | None:
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if s in {"ja", "yes", "true", "1"}:
        return True
    if s in {"nein", "no", "false", "0"}:
        return False
    return None


def _vector(raw: str) -> str:
    return VECTOR_MAP.get(str(raw).strip().lower(), "unknown")


def _malware(name_raw: str) -> list[dict[str, Any]] | None:
    if not name_raw:
        return None
    parts = [p.strip() for p in re.split(r"[,/]| und ", name_raw) if p.strip()]
    cleaned = [re.sub(r"\s*\?$", "", p) for p in parts if p.strip("?")]
    return [{"name": c, "family": None, "mitre_software_id": None} for c in cleaned] or None


def _actor(raw: str) -> tuple[str | None, str]:
    if not raw:
        return None, "unknown"
    s = raw.strip()
    uncertain = s.endswith("?")
    name = s.rstrip("?").strip()
    if name.lower() == "unklar" or not name:
        return None, "unknown"
    return name, ("suspected" if uncertain else "reported")


def _row_to_doc(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    data = dict(zip(headers, row))
    event_date, precision = _parse_date(str(data["Datum"]))
    confirmed = _bool_de(data.get("Bestätigung"))
    actor_name, actor_conf = _actor(str(data.get("Täter") or ""))

    systems = [
        sys_id for col, sys_id in SYSTEM_COLUMNS
        if data.get(col) in (1, "1", True)
    ]

    countries = _countries(str(data.get("Land") or ""))
    slug_seed = (data.get("Name der Malware") or actor_name or "dzsf-import").strip() or "dzsf-import"
    slug = _slugify(str(slug_seed))
    doc_id = f"{event_date.isoformat()}_{slug}"

    doc = {
        "id": doc_id,
        "schema_version": "1.0",
        "event_date": event_date.isoformat(),
        "event_date_precision": precision,
        "reported_date": None,
        "countries_impacted": countries or ["UN"],
        "region": None,
        "coordinates": None,
        "target_organization": None,
        "industry_sector": None,
        "operator_type": None,
        "attack_vector": _vector(str(data.get("Angriffsmittel") or "")),
        "supply_chain": _bool_de(data.get("Supply-Chain")),
        "malware": _malware(str(data.get("Name der Malware") or "")),
        "extortion": _bool_de(data.get("Erpressung")),
        "motive": "financial" if _bool_de(data.get("Erpressung")) else "unknown",
        "mitre_attack_techniques": None,
        "threat_actor": actor_name,
        "threat_actor_country": None,
        "attribution_confidence": actor_conf,
        "confirmation_status": "confirmed" if confirmed else "reported",
        "affected_systems": systems or None,
        "data_published": _bool_de(data.get("Daten veröffentlicht")),
        "impact": None,
        "description_en": (
            f"Imported from DZSF raw data on {event_date.isoformat()}. "
            "Needs review and narrative description."
        ),
        "description_de": None,
        "sources": [
            {
                "url": "https://www.dzsf.bund.de/",
                "type": "academic_paper",
                "publisher": "DZSF",
                "published_date": "2023-05-16",
                "language": "de",
                "accessed_date": None,
            }
        ],
        "contributors": ["DZSF import"],
        "last_updated": date.today().isoformat(),
        "notes": None,
        "provenance": {
            "origin": "DZSF-2020-23-S-1202",
            "imported_from": "20220824_Rohdaten_Cyber_Angriffe.xlsx",
            "needs_review": True,
        },
    }
    return doc


def migrate(xlsx_path: Path, incidents_dir: Path) -> int:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[1])
    count = 0
    for row in rows[2:]:
        if not row or not row[0]:
            continue
        doc = _row_to_doc(headers, row)
        out = incidents_dir / f"{doc['id']}.yaml"
        # Avoid collision: if two incidents share the slug, append index
        n = 2
        while out.exists():
            out = incidents_dir / f"{doc['id']}-{n}.yaml"
            doc["id"] = out.stem
            n += 1
        out.write_text(yaml.safe_dump(doc, sort_keys=False, allow_unicode=True), encoding="utf-8")
        count += 1
    return count


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: migrate_dzsf.py <path/to/Rohdaten.xlsx>", file=sys.stderr)
        return 2
    xlsx = Path(sys.argv[1]).resolve()
    repo_root = Path(__file__).resolve().parent.parent
    count = migrate(xlsx, repo_root / "incidents")
    print(f"Imported {count} incidents into {repo_root/'incidents'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
